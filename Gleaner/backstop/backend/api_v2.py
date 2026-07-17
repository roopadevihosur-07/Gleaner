"""Gleaner platform endpoints — the six-step flow + memory.

connect -> risk -> simulate -> solve -> track (admin) -> memory
"""
from __future__ import annotations

import csv
import io
from typing import Optional

from fastapi import APIRouter, File, UploadFile
from pydantic import BaseModel

import graph as gmod
import store
import tools

router = APIRouter()


# ============================ 1. CONNECT / INGEST ============================
@router.post("/api/connect/upload")
async def connect_upload(file: UploadFile = File(...)):
    """Ingest a company's monitoring export (CSV or XLSX) as live data."""
    raw = await file.read()
    name = (file.filename or "upload").lower()
    columns, rows = [], []

    if name.endswith(".csv") or file.content_type == "text/csv":
        text = raw.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        all_rows = [r for r in reader if any(c.strip() for c in r)]
        if all_rows:
            columns = all_rows[0]
            rows = all_rows[1:]
    elif name.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            data = [[("" if c is None else c) for c in r] for r in ws.iter_rows(values_only=True)]
            data = [r for r in data if any(str(c).strip() for c in r)]
            if data:
                columns = [str(c) for c in data[0]]
                rows = data[1:]
        except Exception as e:
            return {"ok": False, "error": f"Could not parse xlsx: {e}"}
    else:
        return {"ok": False, "error": "Upload a .csv or .xlsx file."}

    preview = [dict(zip(columns, [str(c) for c in r])) for r in rows[:5]]
    meta = store.save_ingested({
        "connected": True, "source": file.filename, "channel": "file",
        "records": len(rows), "columns": columns, "preview": preview,
    })
    return {"ok": True, **meta}


class MCPConnect(BaseModel):
    endpoint: str
    label: Optional[str] = "Company MCP server"


@router.post("/api/connect/mcp")
def connect_mcp(body: MCPConnect):
    """Register a company's MCP endpoint as a live channel. (Handshake is stubbed
    for the demo; the real MCP server is in mcp_server.py.)"""
    meta = store.save_ingested({
        "connected": True, "source": body.endpoint, "channel": "mcp",
        "records": 0, "columns": [], "preview": [], "label": body.label,
    })
    return {"ok": True, **meta}


@router.get("/api/connect/status")
def connect_status():
    return store.get_ingested()


# ============================ 2. RISK CONFIG ============================
class RiskConfig(BaseModel):
    company: Optional[str] = None
    shortfall_trigger_pct: Optional[int] = None
    critical_categories: Optional[list] = None
    auto_approve: Optional[bool] = None
    priority_customers: Optional[list] = None
    objective: Optional[str] = None


@router.get("/api/risk/config")
def get_risk():
    return store.get_risk()


@router.post("/api/risk/config")
def set_risk(cfg: RiskConfig):
    return store.save_risk({k: v for k, v in cfg.dict().items() if v is not None})


# ============================ 3. GRAPH + SIMULATE ============================
@router.get("/api/graph")
def get_graph():
    g = gmod.build_graph()
    return gmod.positioned(g)


class SimulateReq(BaseModel):
    commitment_id: Optional[str] = None
    category: Optional[str] = None
    lost_lbs: Optional[float] = None
    type: Optional[str] = "supply_cut"


@router.post("/api/simulate")
def simulate(req: SimulateReq):
    risk = store.get_risk()
    data = tools.load_all()

    if req.commitment_id:
        gap = tools.size_gap(req.commitment_id)
        category, shortfall = gap["category"], gap["shortfall_lbs"]
        lost = gap["lost_lbs"]
        failed_edge = [req.commitment_id, gmod.DEPOT]
    else:
        category = req.category or "protein"
        lost = req.lost_lbs or 10000
        total_demand = sum(a["demand_lbs"].get(category, 0) for a in data["agencies"])
        on_hand = data["inventory"]["on_hand_lbs"].get(category, 0)
        safety = data["inventory"]["safety_stock_lbs"].get(category, 0)
        shortfall = max(0, total_demand - max(0, on_hand - safety))
        gap = {"category": category, "shortfall_lbs": shortfall, "lost_lbs": lost,
               "total_agency_demand_lbs": total_demand,
               "meals_at_risk": round(tools.lbs_to_meals(shortfall)),
               "affected_agencies": [{"id": a["id"], "name": a["name"], "priority": a["priority"]}
                                     for a in sorted(data["agencies"], key=lambda x: -x["priority"])
                                     if a["demand_lbs"].get(category, 0) > 0]}
        failed_edge = None

    total_demand = gap.get("total_agency_demand_lbs") or 1
    pct = 100 * shortfall / total_demand
    severity = "low"
    if pct >= risk["shortfall_trigger_pct"]:
        severity = "high" if category in risk["critical_categories"] else "medium"
    if category in risk["critical_categories"] and pct >= risk["shortfall_trigger_pct"] * 1.5:
        severity = "critical"

    g = gmod.build_graph()
    viz = gmod.positioned(g, failed_edge=failed_edge)
    solutions = gmod.solutions_for(category, shortfall, risk)
    similar = store.find_similar(category, req.type or "supply_cut")

    return {
        "issue": {
            "id": "ISSUE-" + (req.commitment_id or category).replace(":", "")[:10].upper(),
            "type": req.type or "supply_cut",
            "category": category,
            "lost_lbs": lost,
            "shortfall_lbs": shortfall,
            "shortfall_pct": round(pct, 1),
            "severity": severity,
            "meals_at_risk": gap["meals_at_risk"],
            "affected": gap["affected_agencies"],
            "auto_approve": risk["auto_approve"],
        },
        "graph": viz,
        "solutions": solutions,
        "similar_incidents": similar,
    }


# ============================ 4. SOLVE / DISPATCH ============================
class SolveReq(BaseModel):
    issue: dict
    solution: dict


@router.post("/api/solve")
def solve(req: SolveReq):
    issue, sol = req.issue, req.solution
    category = issue["category"]

    # The routing workaround the local workers execute (divert to protect windows).
    routing = tools.replan_routes(category)
    meals_protected = sol.get("meals_covered", 0) + routing.get("meals_from_perishable", 0)
    dollars = round(routing.get("fuel_saved", 0), 2)

    vendors = ", ".join(v["vendor"] for v in sol.get("vendors", [])) or "n/a"
    protect = ", ".join(a["name"] for a in issue.get("affected", [])[:2])
    worker_instructions = [
        f"Procure {sol.get('filled_lbs', 0):,} lbs {category} from {vendors} "
        f"({sol.get('strategy','')}).",
        f"Divert refrigerated trucks to protect windows at {protect} first.",
        f"Re-sequence route: {routing['baseline']['miles']}\u2192{routing['optimized']['miles']} mi; "
        f"{routing['perishable_lbs_saved']:,} lbs spoilage avoided.",
        "Confirm receipt at depot, then release to delivery.",
    ]

    incident = store.add_incident({
        "type": issue["type"], "category": category, "severity": issue.get("severity"),
        "lost_lbs": issue.get("lost_lbs"), "shortfall_lbs": issue.get("shortfall_lbs"),
        "chosen_strategy": sol.get("strategy"),
        "chosen_vendors": [v["vendor"] for v in sol.get("vendors", [])],
        "worker_instructions": worker_instructions,
        "meals_protected": round(meals_protected),
        "dollars_saved": dollars,
        "status": "resolved" if issue.get("auto_approve") else "dispatched",
    })

    return {
        "incident": incident,
        "dispatch": {"worker_instructions": worker_instructions, "route": routing},
        "summary": {"meals_protected": round(meals_protected), "dollars_saved": dollars},
    }


# ============================ 5. ADMIN / TRACK ============================
@router.get("/api/incidents")
def incidents():
    items = store.list_incidents()
    live = [i for i in items if i["status"] != "resolved"]
    return {
        "incidents": items,
        "metrics": {
            "total": len(items),
            "open": len(live),
            "resolved": len(items) - len(live),
            "meals_protected_total": sum(i.get("meals_protected", 0) for i in items),
        },
    }


class StatusReq(BaseModel):
    status: str  # dispatched | in_progress | resolved


@router.post("/api/incidents/{inc_id}/status")
def set_status(inc_id: str, body: StatusReq):
    rec = store.update_incident(inc_id, {"status": body.status})
    return rec or {"error": "not found"}


# ============================ 6. MEMORY ============================
@router.get("/api/memory/similar")
def memory_similar(category: str, type: str = "supply_cut"):
    return {"similar": store.find_similar(category, type, limit=5)}
